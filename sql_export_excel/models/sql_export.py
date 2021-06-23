# Copyright 2019 Akretion
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html).

import base64
import logging
from io import BytesIO

from odoo import _, api, exceptions, fields, models

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.debug("Can not import openpyxl")


class SqlExport(models.Model):
    _inherit = "sql.export"

    file_format = fields.Selection(
        selection_add=[("excel", "Excel")], ondelete={"excel": "set default"}
    )
    header = fields.Boolean(
        default=True, help="Indicate if the header should be exported to the file."
    )
    attachment_id = fields.Many2one(
        "ir.attachment",
        string="Excel Template",
        help="If you configure an excel file (in xlsx format) here, the "
        "result of the query will be injected in it.\nIt is usefull to "
        "feed data in a excel file pre-configured with calculation",
    )
    sheet_position = fields.Integer(
        default=1,
        help="Indicate the sheet's position of the excel template where the "
        "result of the sql query should be injected.",
    )
    row_position = fields.Integer(
        default=1,
        help="Indicate from which row the result of the query should be " "injected.",
    )
    col_position = fields.Integer(
        string="Column Position",
        default=1,
        help="Indicate from which column the result of the query should be "
        "injected.",
    )

    @api.constrains("sheet_position")
    def check_sheet_position(self):
        for export in self:
            if export.sheet_position < 1:
                raise exceptions.ValidationError(
                    _("The sheet position can't be less than 1.")
                )

    @api.constrains("row_position")
    def check_row_position(self):
        for export in self:
            if export.row_position < 1:
                raise exceptions.ValidationError(
                    _("The row position can't be less than 1.")
                )

    @api.constrains("col_position")
    def check_column_position(self):
        for export in self:
            if export.col_position < 1:
                raise exceptions.ValidationError(
                    _("The column position can't be less than 1.")
                )

    def _get_file_extension(self):
        self.ensure_one()
        if self.file_format == "excel":
            return "xlsx"
        else:
            return super()._get_file_extension()

    def excel_get_data_from_query(self, variable_dict):
        self.ensure_one()
        res = self._execute_sql_request(
            params=variable_dict, mode="fetchall", header=self.header
        )
        # Case we insert data in an existing excel file.
        if self.attachment_id:
            datas = self.attachment_id.datas
            infile = BytesIO()
            infile.write(base64.b64decode(datas))
            infile.seek(0)
            wb = openpyxl.load_workbook(filename=infile)
            sheets = wb.worksheets
            try:
                ws = sheets[self.sheet_position - 1]
            except IndexError:
                raise exceptions.ValidationError(
                    _(
                        "The Excel Template file contains less than %s sheets "
                        "Please, adjust the Sheet Position parameter."
                    )
                )
            row_position = self.row_position or 1
            col_position = self.col_position or 1
        # Case of excel file creation
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            row_position = 1
            col_position = 1
        for index, row in enumerate(res, row_position):
            for col, val in enumerate(row, col_position):
                ws.cell(row=index, column=col).value = val
        output = BytesIO()
        wb.save(output)
        output.getvalue()
        output_datas = base64.b64encode(output.getvalue())
        output.close()
        return output_datas
