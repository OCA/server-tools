# Copyright (C) 2019 Akretion (<http://www.akretion.com>)
# @author: Florian da Costa
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).

import base64
import logging
from io import BytesIO

from odoo.tests.common import TransactionCase

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    _logger.debug("Can not import openpyxl")


class TestExportSqlQueryExcel(TransactionCase):
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.wizard_obj = cls.env["sql.file.wizard"]

    def get_workbook_from_query(self, wizard):
        wizard.export_sql()
        decoded_data = base64.b64decode(wizard.binary_file)
        xlsx_file = BytesIO(decoded_data)
        return openpyxl.load_workbook(xlsx_file)

    def test_excel_file_generation(self):
        test_query = "SELECT 'testcol1' as firstcol, 2 as second_col"
        query_vals = {
            "name": "Test Query Excel",
            "query": test_query,
            "file_format": "excel",
        }
        query = self.env["sql.export"].create(query_vals)
        query.button_validate_sql_expression()
        wizard = self.wizard_obj.create(
            {
                "sql_export_id": query.id,
            }
        )
        workbook = self.get_workbook_from_query(wizard)
        ws = workbook.active
        # Check values, header should be here by default
        self.assertEqual(ws.cell(row=1, column=1).value, "firstcol")
        self.assertEqual(ws.cell(row=2, column=1).value, "testcol1")
        self.assertEqual(ws.cell(row=2, column=2).value, 2)

        query.write({"header": False})
        wb2 = self.get_workbook_from_query(wizard)
        ws2 = wb2.active
        # Check values, the header should not be present
        self.assertEqual(ws2.cell(row=1, column=1).value, "testcol1")
        self.assertEqual(ws2.cell(row=1, column=2).value, 2)

    def test_excel_file_insert(self):
        # Create excel file with 2 sheets. Create a header in second sheet
        # where data will be inserted
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="My Test Value")
        ws2 = wb.create_sheet("data")
        ws2.cell(row=1, column=1, value="Partner Id")
        ws2.cell(row=1, column=2, value="Partner Name")
        output = BytesIO()
        wb.save(output)
        data = output.getvalue()

        # Create attachment with the created xlsx file which will be used as
        # template in the sql query
        attachmnent_vals = {
            "name": "template xlsx sql export Res Partner",
            "datas": base64.b64encode(data),
        }
        attachment = self.env["ir.attachment"].create(attachmnent_vals)

        # Create the query and configure it to insert the data in the second
        # sheet of the xlsx template file and start inserting data at the
        # second row, ignoring header (because the template excel file
        # already contains a header)
        test_query = "SELECT id, name FROM res_partner"
        query_vals = {
            "name": "Test Query Excel",
            "query": test_query,
            "file_format": "excel",
            "attachment_id": attachment.id,
            "sheet_position": 2,
            "header": False,
            "row_position": 2,
        }
        query = self.env["sql.export"].create(query_vals)
        query.button_validate_sql_expression()
        wizard = self.wizard_obj.create(
            {
                "sql_export_id": query.id,
            }
        )

        # Check the generated excel file. The first sheet should still contain
        # the same data and the second sheet should have kept the header and
        # inserted data from the query
        wb2 = self.get_workbook_from_query(wizard)
        sheets = wb2.worksheets
        ws1 = sheets[0]
        # Check values, header should be here by default
        self.assertEqual(ws1.cell(row=1, column=1).value, "My Test Value")
        ws2 = sheets[1]
        self.assertEqual(ws2.cell(row=1, column=1).value, "Partner Id")
        self.assertTrue(ws2.cell(row=2, column=1).value)
