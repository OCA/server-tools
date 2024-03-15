# © 2022 initOS GmbH
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
import csv
import io
import json
import logging
import tempfile
import uuid
import zipfile
from itertools import tee

import chardet
from lxml import etree
from pytz import timezone

from odoo import _, api, fields, models
from odoo.exceptions import UserError
from odoo.tools import safe_eval

_logger = logging.getLogger(__name__)

try:
    # Optional feature for xlsx
    import openpyxl
except ImportError:
    _logger.warning(
        "The openpyxl python library is not installed. The XLSX feature isn't available"
    )
    openpyxl = None


try:
    # Optional feature for JSONPath
    import jsonpath_ng as jsonpath
except ImportError:
    _logger.warning(
        "The jsonpath-ng python library is not installed. "
        "The JSONPath feature isn't available"
    )
    jsonpath = None


re = safe_eval.wrap_module(
    __import__("re"),
    [
        "compile",
        "escape",
        "findall",
        "finditer",
        "fullmatch",
        "match",
        "search",
        "sub",
        "subn",
    ],
)


def chunking(items, size):
    if size < 1:
        yield from items
        return

    chunk = []
    for item in items:
        chunk.append(item)
        if len(chunk) >= size:
            yield chunk
            chunk = []

    if chunk:
        yield chunk


class ImportProcessor(models.Model):
    _name = "import.processor"
    _description = _("Generic Import Processor")
    _order = "name"

    def _get_file_types(self):
        return [
            ("csv", "CSV"),
            ("json", "JSON"),
            ("xlsx", "XLSX"),
            ("xml", "XML"),
        ]

    def _get_csv_delimiter(self):
        return [
            ("comma", _("Comma")),
            ("semicolon", _("Semicolon")),
            ("tab", _("Tab")),
        ]

    def _get_compression(self):
        return [
            ("zip_one", _("Zipped File")),
            ("zip_all", _("Multiple Zipped Files")),
        ]

    def _get_default_code(self, entry=False):
        variables = self.default_variables()
        if not entry:
            for key in ("entry", "key"):
                variables.pop(key, None)

        desc = "\n".join(f"# - {v}: {desc}" for v, desc in variables.items())
        return f"# Possible variables:\n{desc}\n\n"

    name = fields.Char(translate=True, required=True)
    active = fields.Boolean(default=True)
    model_id = fields.Many2one(
        "ir.model",
        string="Import Model",
        ondelete="cascade",
        required=True,
    )
    model_name = fields.Char(related="model_id.model", related_sudo=True)
    file_type = fields.Selection(_get_file_types, default="csv", required=True)
    processor = fields.Text(default=lambda self: self._get_default_code(True))
    preprocessor = fields.Text(default=_get_default_code)
    postprocessor = fields.Text(default=_get_default_code)
    help_text = fields.Html(compute="_compute_help_text", readonly=True, store=False)
    file_encoding = fields.Char()
    compression = fields.Selection(_get_compression)
    chunk_size = fields.Integer(
        default=0,
        help="If greater than 0 it script will get a list of entries instead to "
        "speed up the processing up. Chunking is only supported for streamable data "
        "line csv or xlsx",
    )

    csv_delimiter = fields.Selection(
        _get_csv_delimiter,
        "CSV Delimiter",
        default="comma",
    )
    csv_row_offset = fields.Integer("Row Offset", default=0)
    csv_quotechar = fields.Char("Quote Char", default='"', size=1)

    json_path_entry = fields.Char("JSONPath Entry")

    xml_path_entry = fields.Char("XPath Entry")
    xml_namespaces = fields.Text("XML Namespace")

    tabular_sheet_name = fields.Char("Sheet Name")
    tabular_sheet_index = fields.Integer("Sheet Index", default=0)
    tabular_col_offset = fields.Integer("Sheet Col Offset", default=0)
    tabular_row_offset = fields.Integer("Sheet Row Offset", default=0)

    @api.model
    def _name_search(
        self, name="", args=None, operator="ilike", limit=100, name_get_uid=None
    ):
        return super(ImportProcessor, self.sudo())._name_search(
            name=name,
            args=args,
            operator=operator,
            limit=limit,
            name_get_uid=name_get_uid,
        )

    def _compute_help_text(self):
        variables = self.default_variables()
        lines = []
        for var, desc in variables.items():
            var = (f"<code>{v.strip()}</code>" for v in var.split(","))
            lines.append(f"<li>{', '.join(sorted(var))}: {desc}</li>")

        desc = "\n".join(lines)
        self.write({"help_text": f"<ul>{desc}</ul>"})

    def _get_eval_context(self):
        self.ensure_one()

        def log(message, *args, level="info"):
            level = {
                "debug": logging.DEBUG,
                "info": logging.INFO,
                "warning": logging.WARNING,
                "error": logging.ERROR,
                "critical": logging.CRITICAL,
            }.get(level, logging.INFO)
            _logger.log(level, f"{self} | {message}", *args)

        return {
            "datetime": safe_eval.datetime,
            "env": self.env,
            "log": log,
            "model": self._get_model(),
            "re": re,
            "time": safe_eval.time,
            "timezone": timezone,
            "UserError": UserError,
        }

    def _csv_get_delimiter(self):
        self.ensure_one()
        return {
            "comma": ",",
            "semicolon": ";",
            "tab": "\t",
        }.get(self.csv_delimiter, ",")

    def _get_model(self):
        self.ensure_one()
        return self.env[self.sudo().model_id.model]

    def default_variables(self):
        """Informations about the available variables in the python code"""
        return {
            "entry": "The data entry to import into odoo",
            "env": "Odoo Environment on which the import is triggered",
            "header": "Will be used as header instead of reading it if CSV",
            "key": "The key/index of the entry if JSON",
            "log": "Logging functions",
            "model": "Odoo Model of the record on which the action is triggered",
            "nsmap": "The XML namespaces if xml",
            "process_uuid": "Unique UUID for each import process",
            "reader": "CSV reader if CSV. The importing continues",
            "records": "The already imported records",
            "root": "XML root if XML or JSON",
            "sheet": "Tabular sheet if XLSX",
            "datetime, re, time, timezone": "useful Python libraries",
            "UserError": "Warning Exception to use with raise",
        }

    def _process_entry(self, process_uuid, entry, localdict, **kwargs):
        # Reset the pre-defined values
        localdict.update(self._get_eval_context())
        localdict.update({"entry": entry, "process_uuid": process_uuid, **kwargs})

        safe_eval.safe_eval(self.processor, localdict, mode="exec", nocopy=True)
        localdict.pop("entry", None)
        for key in kwargs:
            localdict.pop(key, None)

    def process_entry(self, process_uuid, entry, localdict, **kwargs):
        # Reset the pre-defined values
        self._process_entry(process_uuid, entry, localdict, **kwargs)

    def _pre_process(self, process_uuid, localdict):
        if self.preprocessor:
            localdict["process_uuid"] = process_uuid
            localdict.update(self._get_eval_context())
            safe_eval.safe_eval(self.preprocessor, localdict, mode="exec", nocopy=True)

    def _post_process(self, process_uuid, localdict):
        # Reset the pre-defined values
        if self.postprocessor:
            localdict["process_uuid"] = process_uuid
            localdict.update(self._get_eval_context())
            safe_eval.safe_eval(self.postprocessor, localdict, mode="exec", nocopy=True)

    def _process_csv(self, process_uuid, file):
        if isinstance(file, bytes):
            encoding = self.file_encoding or chardet.detect(file)["encoding"] or "utf-8"
            file = io.TextIOWrapper(io.BytesIO(file), encoding=encoding.lower())
        elif isinstance(file, str):
            encoding = self.file_encoding or "utf-8"
            file = io.TextIOWrapper(io.StringIO(file), encoding=encoding)

        reader = csv.reader(
            file,
            delimiter=self._csv_get_delimiter(),
            quotechar=self.csv_quotechar or '"',
        )
        reader, backup = tee(reader)
        localdict = {"records": self._get_model(), "reader": reader}

        # Apply the pre-processor to initialize the environment
        self._pre_process(process_uuid, localdict)

        localdict.pop("reader", None)
        # The next line is the CSV field header. Afterwards parse every line
        for _x in range(self.csv_row_offset):
            next(backup)

        header = localdict.pop("header", None)
        if header is None:
            header = next(backup)

        def prepare_line(line):
            return dict(list(zip(header, line))[::-1])

        for data in chunking(map(prepare_line, backup), self.chunk_size):
            self.process_entry(process_uuid, data, localdict)

        # Do some final post processing
        self._post_process(process_uuid, localdict)
        return localdict.get("records", self._get_model())

    def _process_json(self, process_uuid, file):
        if isinstance(file, bytes):
            encoding = self.file_encoding or chardet.detect(file)["encoding"] or "utf-8"
            file = io.TextIOWrapper(io.BytesIO(file), encoding=encoding.lower())
        elif isinstance(file, str):
            encoding = self.file_encoding or "utf-8"
            file = io.TextIOWrapper(io.StringIO(file), encoding=encoding)

        root = json.load(file)
        localdict = {"records": self._get_model(), "root": root}

        # Apply the pre-processor to initialize the environment
        self._pre_process(process_uuid, localdict)

        # Check and apply the JSONpath if module is available
        if self.json_path_entry and not jsonpath:
            raise UserError(
                _(
                    "The JSONPath isn't available because the module jsonpath-ng is "
                    "missing. Please contact your administrator"
                )
            )

        if self.json_path_entry:
            found = jsonpath.parse(self.json_path_entry).find(root)
            if all(isinstance(x.path, jsonpath.Index) for x in found):
                root = [x.value for x in found]
            elif all(isinstance(x.path, jsonpath.Fields) for x in found):
                root = {str(x.path): x.value for x in found}
            else:
                raise UserError(_("Unexpected JSON file"))

        if isinstance(root, dict):
            iterator = root.items()
        else:
            iterator = enumerate(root)

        # Iterate over the entries
        for key, entry in iterator:
            self.process_entry(process_uuid, entry, localdict, key=key)

        # Do some final post processing
        self._post_process(process_uuid, localdict)
        return localdict.get("records", self._get_model())

    def _process_xml(self, process_uuid, file):
        root = etree.fromstring(file)

        # Get the namespace
        if self.xml_namespaces:
            nsmap = safe_eval.safe_eval(self.xml_namespaces)
        else:
            nsmap = None

        localdict = {"records": self._get_model(), "root": root, "nsmap": nsmap}

        # Apply the pre-processor to initialize the environment
        self._pre_process(process_uuid, localdict)

        for entry in root.xpath(self.xml_path_entry, namespaces=nsmap):
            self.process_entry(process_uuid, entry, localdict)

        # Do some final post processing
        self._post_process(process_uuid, localdict)
        return localdict.get("records", self._get_model())

    def _process_xlsx(self, process_uuid, file):
        if not openpyxl:
            raise UserError(
                _(
                    "The XLSX processing isn't available because openpyxl is "
                    "missing. Please contact your administrator"
                )
            )

        if isinstance(file, bytes):
            with tempfile.NamedTemporaryFile("wb", suffix=".xlsx") as fp:
                fp.write(file)
                fp.flush()
                workbook = openpyxl.load_workbook(fp.name)
        elif isinstance(file, str):
            with tempfile.NamedTemporaryFile("w", suffix=".xlsx") as fp:
                fp.write(file)
                fp.flush()

                workbook = openpyxl.load_workbook(fp.name)

        if self.tabular_sheet_name:
            sheet = workbook[self.tabular_sheet_name]
        else:
            sheet_name = workbook.sheetnames[self.tabular_sheet_index]
            sheet = workbook[sheet_name]

        localdict = {"records": self._get_model(), "sheet": sheet}

        # Apply the pre-processor to initialize the environment
        self._pre_process(process_uuid, localdict)

        localdict.pop("workbook", None)

        reader = sheet.rows
        for _x in range(self.tabular_row_offset):
            next(reader)

        def prepare_line(line):
            entry = [(h, cell.value) for h, cell in zip(header, line[col:])]
            return dict(entry[::-1])

        # The next line is the field header. Afterwards parse every line
        col = self.tabular_col_offset
        header = [x.value for x in next(reader)[col:]]
        for data in chunking(map(prepare_line, reader), self.chunk_size):
            self.process_entry(process_uuid, data, localdict)

        # Do some final post processing
        self._post_process(process_uuid, localdict)
        return localdict.get("records", self._get_model())

    def process(self, file):
        self.ensure_one()

        process_uuid = str(uuid.uuid4())

        method = getattr(self, f"_process_{self.file_type}", None)
        if not callable(method):
            raise NotImplementedError()

        if self.compression in ("zip_one", "zip_all"):
            with tempfile.NamedTemporaryFile("wb+") as fp:
                fp.write(file)
                fp.flush()

                try:
                    zipped = zipfile.ZipFile(fp.name)
                except zipfile.BadZipfile:
                    _logger.warning("File is no zip. Falling back: %s", self)
                    return method(process_uuid, file)

            if self.compression == "zip_one" and len(zipped.filelist) != 1:
                raise UserError(_("Expected only 1 file."))

            result = self._get_model()
            for zipped_file in zipped.filelist:
                result |= method(process_uuid, zipped.read(zipped_file))
            return result

        return method(process_uuid, file)
