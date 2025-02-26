# Copyright 2019 Ecosoft Co., Ltd (http://ecosoft.co.th/)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl.html)

import base64
import uuid
from ast import literal_eval
from datetime import date
from datetime import datetime as dt
from io import BytesIO

import xlrd
import xlwt

from odoo import api, models
from odoo.exceptions import UserError, ValidationError
from odoo.tools.float_utils import float_compare
from odoo.tools.parse_version import parse_version
from odoo.tools.safe_eval import safe_eval

from . import common as co


class XLSXImport(models.AbstractModel):
    _name = "xlsx.import"
    _description = "Excel Import AbstractModel"

    @api.model
    def get_eval_context(self, model=False, value=False):
        eval_context = {
            "float_compare": float_compare,
            "datetime": dt,
            "date": date,
            "env": self.env,
            "context": self._context,
            "value": False,
            "model": False,
        }
        if model:
            eval_context.update({"model": self.env[model]})
        if value:
            if isinstance(value, str):  # Remove non Ord 128 character
                value = "".join([i if ord(i) < 128 else " " for i in value])
            eval_context.update({"value": value})
        return eval_context

    @api.model
    def get_external_id(self, record):
        """Get external ID of the record, if not already exists create one"""
        ModelData = self.sudo().env["ir.model.data"]
        xml_id = record.get_external_id()
        if not xml_id or (record.id in xml_id and xml_id[record.id] == ""):
            ModelData.create(
                {
                    "name": f"{record._table}_{record.id}",
                    "module": "__excel_import_export__",
                    "model": record._name,
                    "res_id": record.id,
                }
            )
            xml_id = record.get_external_id()
        return xml_id[record.id]

    @api.model
    def _get_field_type(self, model, field):
        try:
            record = self.env[model].new()
            for f in field.split("/"):
                field_type = record._fields[f].type
                if field_type in ("one2many", "many2many"):
                    record = record[f]
                else:
                    return field_type
        except Exception as exc:
            raise ValidationError(
                self.env._("Invalid declaration, %s has no valid field type", field)
            ) from exc

    @api.model
    def _delete_record_data(self, record, data_dict):
        """If no _NODEL_, delete existing lines before importing"""
        if not record or not data_dict:
            return
        try:
            for sheet_name in data_dict:
                worksheet = data_dict[sheet_name]
                line_fields = filter(lambda x: x != "_HEAD_", worksheet)
                for line_field in line_fields:
                    if "_NODEL_" not in line_field:
                        if line_field in record and record[line_field]:
                            record[line_field].unlink()
            # Remove _NODEL_ from dict
            for s, _sv in data_dict.copy().items():
                for f, _fv in data_dict[s].copy().items():
                    if "_NODEL_" in f:
                        new_fv = data_dict[s].pop(f)
                        data_dict[s][f.replace("_NODEL_", "")] = new_fv
        except Exception as e:
            raise ValidationError(self.env._("Error deleting data\n%s", e)) from e

    @api.model
    def _get_end_row(self, st, worksheet, line_field, is_xlsx=False):
        """Get max row or next empty row as the ending row"""
        _x, max_row = co.get_line_max(line_field)
        test_rows = {}
        max_end_row = 0
        if is_xlsx:
            # For openpyxl
            sheet_nrows = st.max_row
        else:
            # For xlrd
            sheet_nrows = st.nrows
        for rc, _col in worksheet.get(line_field, {}).items():
            rc, key_eval_cond = co.get_field_condition(rc)
            row, col = co.pos2idx(rc)
            # Use max_row, i.e., order_line[5], use it. Otherwise, use st.nrows
            max_end_row = sheet_nrows if max_row is False else (row + max_row)
            for idx in range(row, max_row and max_end_row or sheet_nrows):
                try:
                    if is_xlsx:
                        # openpyxl uses 1-based indexing
                        cell = st.cell(row=idx + 1, column=col + 1)
                        # Check if cell is empty (None or empty string)
                        cell_type = 0 if cell.value in (None, "") else 1
                    else:
                        cell_type = st.cell_type(idx, col)  # empty type = 0
                except Exception as e:
                    raise UserError(
                        self.env._(
                            "The value for the '%(field)s' field is expected to be "
                            "in cell %(cell_position)s, but no column exists for that "
                            "cell in the Excel sheet. Please check your Excel file.",
                            field=_col,
                            cell_position=rc,
                        )
                    ) from e
                r_types = test_rows.get(idx, [])
                r_types.append(cell_type)
                test_rows[idx] = r_types
        empty_list = filter(lambda y: all(i == 0 for i in y[1]), test_rows.items())
        empty_rows = list(map(lambda z: z[0], empty_list))
        next_empty_row = empty_rows and min(empty_rows) or max_end_row
        return next_empty_row

    @api.model
    def _get_line_vals(self, st, worksheet, model, line_field, is_xlsx=False):
        """Get values of this field from excel sheet"""
        vals = {}
        end_row = self._get_end_row(st, worksheet, line_field, is_xlsx)
        for rc, columns in worksheet.get(line_field, {}).items():
            if not isinstance(columns, list):
                columns = [columns]
            for field in columns:
                rc, key_eval_cond = co.get_field_condition(rc)
                x_field, val_eval_cond = co.get_field_condition(field)
                row, col = co.pos2idx(rc)
                new_line_field, _x = co.get_line_max(line_field)
                out_field = f"{new_line_field}/{x_field}"
                field_type = self._get_field_type(model, out_field)
                vals.update({out_field: []})
                for idx in range(row, end_row):
                    if is_xlsx:  # Handle openpyxl cells
                        cell_value = st.cell(row=idx + 1, column=col + 1)
                    else:  # Handle xlrd cells
                        cell_value = st.cell(idx, col)
                    value = co._get_cell_value(cell_value, field_type=field_type)
                    eval_context = self.get_eval_context(model=model, value=value)
                    if key_eval_cond:
                        value = safe_eval(key_eval_cond, eval_context)
                    if val_eval_cond:
                        value = safe_eval(val_eval_cond, eval_context)
                    vals[out_field].append(value)
                if not filter(lambda x: x != "", vals[out_field]):
                    vals.pop(out_field)
        return vals

    @api.model
    def _process_worksheet(
        self, wb, out_st, model, data_dict, header_fields, is_xlsx=False
    ):  # noqa: C901
        col_idx = 1
        for sheet_name in data_dict:  # For each Sheet
            worksheet = data_dict[sheet_name]
            st = None
            if isinstance(sheet_name, str):
                # Get sheet by name for openpyxl
                st = (
                    wb[sheet_name]
                    if is_xlsx
                    else co.xlrd_get_sheet_by_name(wb, sheet_name)
                )
            elif isinstance(sheet_name, int):
                # Get sheet by index for openpyxl
                st = (
                    wb.worksheets[sheet_name - 1]
                    if is_xlsx
                    else wb.sheet_by_index(sheet_name - 1)
                )
            if not st:
                raise ValidationError(self.env._("Sheet %s not found", sheet_name))
            # HEAD updates
            for rc, field in worksheet.get("_HEAD_", {}).items():
                rc, key_eval_cond = co.get_field_condition(rc)
                field, val_eval_cond = co.get_field_condition(field)
                field_type = self._get_field_type(model, field)
                try:
                    row, col = co.pos2idx(rc)
                    if is_xlsx:  # openpyxl
                        # openpyxl uses 1-based indexing
                        cell_value = st.cell(row=row + 1, column=col + 1).value
                    else:  # xlrd
                        cell_value = st.cell(row, col)
                    value = co._get_cell_value(cell_value, field_type=field_type)
                except Exception:
                    value = False
                eval_context = self.get_eval_context(model=model, value=value)
                if key_eval_cond:
                    value = str(safe_eval(key_eval_cond, eval_context))
                if val_eval_cond:
                    value = str(safe_eval(val_eval_cond, eval_context))
                out_st.write(0, col_idx, field)  # Next Column
                out_st.write(1, col_idx, value)  # Next Value
                header_fields.append(field)
                col_idx += 1
            # Line Items
            line_fields = filter(lambda x: x != "_HEAD_", worksheet)
            for line_field in line_fields:
                vals = self._get_line_vals(st, worksheet, model, line_field, is_xlsx)
                for field in vals:
                    # Columns, i.e., line_ids/field_id
                    out_st.write(0, col_idx, field)
                    header_fields.append(field)
                    # Data
                    i = 1
                    for value in vals[field]:
                        out_st.write(i, col_idx, value)
                        i += 1
                    col_idx += 1

    @api.model
    def _import_record_data(self, import_file, record, data_dict):
        if not data_dict:
            return
        try:
            model = record._name
            xml_id = self._generate_xml_id(record)
            decoded_data = base64.decodebytes(import_file)
            out_wb = xlwt.Workbook()
            out_st = out_wb.add_sheet("Sheet 1")
            out_st.write(0, 0, "id")
            out_st.write(1, 0, xml_id)
            header_fields = ["id"]
            is_xlsx = self._is_xlsx_file()
            wb = self._load_workbook(decoded_data, is_xlsx)
            # Process on all worksheets
            self._process_worksheet(
                wb, out_st, model, data_dict, header_fields, is_xlsx
            )
            content = BytesIO()
            out_wb.save(content)
            content.seek(0)  # Set index to 0, and start reading
            xls_file = content.read()
            # Do the import
            imp = self._create_import_record(model, xls_file)
            errors = self._execute_import(imp, header_fields)
            if errors.get("messages"):
                self._handle_import_errors(errors)
            return self.env.ref(xml_id)
        except xlrd.XLRDError as exc:
            raise ValidationError(
                self.env._("Invalid file style, only .xls or .xlsx file allowed")
            ) from exc
        except Exception as e:
            raise e

    def _generate_xml_id(self, record):
        return (
            record
            and self.get_external_id(record)
            or "{}.{}".format("__excel_import_export__", uuid.uuid4())
        )

    def _is_xlsx_file(self):
        if xlrd and parse_version(xlrd.__VERSION__) < parse_version("2.0"):
            return False
        return True

    def _load_workbook(self, decoded_data, is_xlsx):
        if not is_xlsx:
            return xlrd.open_workbook(file_contents=decoded_data)
        try:
            import io

            from openpyxl import load_workbook

            return load_workbook(io.BytesIO(decoded_data or b""), data_only=True)
        except Exception as exc:
            raise ValidationError(
                self.env._("Invalid file style, only .xls or .xlsx file allowed")
            ) from exc

    def _create_import_record(self, model, xls_file):
        Import = self.env["base_import.import"]
        return Import.create(
            {
                "res_model": model,
                "file": xls_file,
                "file_type": "application/vnd.ms-excel",
                "file_name": "temp.xls",
            }
        )

    def _execute_import(self, imp, header_fields):
        return imp.execute_import(
            header_fields,
            header_fields,
            {
                "has_headers": True,
                "advanced": True,
                "keep_matches": False,
                "encoding": "",
                "separator": "",
                "quoting": '"',
                "date_format": "%Y-%m-%d",
                "datetime_format": "%Y-%m-%d %H:%M:%S",
                "float_thousand_separator": ",",
                "float_decimal_separator": ".",
                "fields": [],
            },
        )

    def _handle_import_errors(self, errors):
        message = self.env._("Error importing data")
        messages = errors["messages"]
        if isinstance(messages, dict):
            message = messages["message"]
        if isinstance(messages, list):
            message = ", ".join([x["message"] for x in messages])
        raise ValidationError(message.encode("utf-8"))

    @api.model
    def _post_import_operation(self, record, operation):
        """Run python code after import"""
        if not record or not operation:
            return
        try:
            if "${" in operation:
                code = (operation.split("${"))[1].split("}")[0]
                eval_context = {"object": record}
                safe_eval(code, eval_context)
        except Exception as e:
            raise ValidationError(
                self.env._("Post import operation error\n%s", e)
            ) from e

    @api.model
    def import_xlsx(self, import_file, template, res_model=False, res_id=False):
        """
        - If res_id = False, we want to create new document first
        - Delete fields' data according to data_dict['__IMPORT__']
        - Import data from excel according to data_dict['__IMPORT__']
        """
        if res_model and template.res_model != res_model:
            raise ValidationError(self.env._("Template's model mismatch"))
        record = self.env[template.res_model].browse(res_id)
        data_dict = literal_eval(template.instruction.strip())
        if not data_dict.get("__IMPORT__"):
            raise ValidationError(
                self.env._("No data_dict['__IMPORT__'] in template %s", template.name)
            )
        if record:
            # Delete existing data first
            self._delete_record_data(record, data_dict["__IMPORT__"])
        # Fill up record with data from excel sheets
        record = self._import_record_data(import_file, record, data_dict["__IMPORT__"])
        # Post Import Operation, i.e., cleanup some data
        if data_dict.get("__POST_IMPORT__", False):
            self._post_import_operation(record, data_dict["__POST_IMPORT__"])
        return record
